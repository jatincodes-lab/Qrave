import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Error, sync_playwright


DEFAULT_OUTPUT = "bareilly_qrave_restaurant_leads.xlsx"

HEADERS = [
    "captured_at",
    "restaurant_name",
    "phone_numbers",
    "emails",
    "address_guess",
    "source_url",
    "source_title",
    "website_links",
    "google_maps_links",
    "social_links",
    "menu_links",
    "ordering_links",
    "has_menu_signal",
    "has_ordering_signal",
    "qrave_fit_score",
    "qrave_fit_reason",
    "notes",
    "page_text_preview",
]

PHONE_RE = re.compile(
    r"(?:\+91[\s-]?)?(?:0[\s-]?)?[6-9]\d{2}[\s-]?\d{3}[\s-]?\d{4}"
)
EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)

RESTAURANT_WORDS = (
    "restaurant",
    "cafe",
    "coffee",
    "dhaba",
    "biryani",
    "pizza",
    "burger",
    "bakery",
    "kitchen",
    "food",
    "dining",
    "sweets",
    "fast food",
)

ADDRESS_WORDS = (
    "bareilly",
    "uttar pradesh",
    "u.p.",
    "up ",
    "road",
    "nagar",
    "colony",
    "market",
    "near",
    "opposite",
    "civil lines",
    "dd puram",
    "rajendra nagar",
    "pilibhit bypass",
)

MENU_WORDS = ("menu", "catalog", "price list", "rate list")
ORDERING_WORDS = (
    "order online",
    "online order",
    "delivery",
    "takeaway",
    "take away",
    "swiggy",
    "zomato",
    "magicpin",
    "dotpe",
    "petpooja",
    "posist",
    "urbanpiper",
    "wa.me",
    "api.whatsapp.com",
)
SOCIAL_DOMAINS = (
    "instagram.com",
    "facebook.com",
    "fb.com",
    "youtube.com",
    "wa.me",
    "api.whatsapp.com",
)
MAPS_DOMAINS = ("google.com/maps", "maps.app.goo.gl", "goo.gl/maps")


def clean_space(value):
    return re.sub(r"\s+", " ", value or "").strip()


def unique(values):
    seen = set()
    result = []
    for value in values:
        if not value:
            continue
        key = value.strip().lower()
        if key and key not in seen:
            seen.add(key)
            result.append(value.strip())
    return result


def normalize_phone(value):
    digits = re.sub(r"\D", "", value)
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    if digits.startswith("0") and len(digits) == 11:
        digits = digits[1:]
    if len(digits) == 10:
        return digits
    return ""


def host(url):
    try:
        return urlparse(url).netloc.lower().replace("www.", "")
    except ValueError:
        return ""


def extract_jsonld_candidates(jsonld_items):
    candidates = []
    for item in jsonld_items:
        parsed_items = item if isinstance(item, list) else [item]
        for parsed in parsed_items:
            if isinstance(parsed, dict) and "@graph" in parsed:
                graph = parsed.get("@graph") or []
                parsed_items.extend(graph if isinstance(graph, list) else [graph])
                continue
            if isinstance(parsed, dict):
                candidates.append(parsed)
    return candidates


def format_address(address):
    if isinstance(address, str):
        return clean_space(address)
    if not isinstance(address, dict):
        return ""
    parts = [
        address.get("streetAddress"),
        address.get("addressLocality"),
        address.get("addressRegion"),
        address.get("postalCode"),
        address.get("addressCountry"),
    ]
    return clean_space(", ".join(str(part) for part in parts if part))


def find_restaurant_schema(jsonld_candidates):
    for item in jsonld_candidates:
        item_type = item.get("@type", "")
        types = item_type if isinstance(item_type, list) else [item_type]
        normalized_types = {str(value).lower() for value in types}
        if normalized_types.intersection(
            {"restaurant", "cafeorcoffeeshop", "foodestablishment", "localbusiness"}
        ):
            return item
    return {}


def choose_name(schema, title, meta):
    name = schema.get("name") if isinstance(schema, dict) else ""
    if isinstance(name, dict):
        name = name.get("text", "")
    if name:
        return clean_space(str(name))

    for key in ("og_site_name", "og_title", "meta_title"):
        if meta.get(key):
            return clean_space(meta[key])

    title = re.split(r"\s[-|]\s", title or "")[0]
    return clean_space(title)


def guess_address(text, schema):
    schema_address = format_address(schema.get("address")) if isinstance(schema, dict) else ""
    if schema_address:
        return schema_address

    lines = [clean_space(line) for line in text.splitlines()]
    candidates = []
    for line in lines:
        lowered = line.lower()
        if 15 <= len(line) <= 180 and any(word in lowered for word in ADDRESS_WORDS):
            candidates.append(line)
    return " | ".join(unique(candidates[:3]))


def score_lead(text, phones, address, menu_links, ordering_links, current_url):
    lowered = text.lower()
    score = 0
    reasons = []

    if phones:
        score += 20
        reasons.append("phone found")
    if "bareilly" in lowered or "bareilly" in address.lower():
        score += 20
        reasons.append("Bareilly signal")
    if any(word in lowered for word in RESTAURANT_WORDS):
        score += 20
        reasons.append("restaurant/cafe signal")
    if menu_links:
        score += 10
        reasons.append("menu signal")
    else:
        score += 15
        reasons.append("no menu link found")
    if ordering_links:
        score -= 10
        reasons.append("already has ordering signal")
    else:
        score += 20
        reasons.append("no ordering link found")
    if host(current_url) in {"google.com", "maps.google.com"}:
        score -= 10
        reasons.append("captured from Google page")

    return max(0, min(score, 100)), "; ".join(reasons)


def extract_from_page(page):
    payload = page.evaluate(
        """
        () => {
          const meta = {};
          for (const element of document.querySelectorAll('meta')) {
            const key = element.getAttribute('property') || element.getAttribute('name');
            if (!key) continue;
            meta[key] = element.getAttribute('content') || '';
          }

          const jsonld = [];
          for (const script of document.querySelectorAll('script[type="application/ld+json"]')) {
            const raw = script.textContent || '';
            try {
              jsonld.push(JSON.parse(raw));
            } catch (_) {}
          }

          const links = Array.from(document.querySelectorAll('a')).map((a) => ({
            text: (a.innerText || a.getAttribute('aria-label') || '').trim(),
            href: a.href || ''
          })).filter((link) => link.href);

          return {
            title: document.title || '',
            url: location.href,
            text: document.body ? document.body.innerText : '',
            meta,
            jsonld,
            links
          };
        }
        """
    )

    text = payload.get("text") or ""
    links = payload.get("links") or []
    jsonld_candidates = extract_jsonld_candidates(payload.get("jsonld") or [])
    schema = find_restaurant_schema(jsonld_candidates)

    schema_phone = schema.get("telephone", "") if isinstance(schema, dict) else ""
    phones = [normalize_phone(match.group(0)) for match in PHONE_RE.finditer(text)]
    phones.extend(normalize_phone(value) for value in re.split(r"[,/|]", str(schema_phone)))
    phones = unique([phone for phone in phones if phone])

    emails = unique(match.group(0) for match in EMAIL_RE.finditer(text))

    meta = {
        "og_title": payload.get("meta", {}).get("og:title", ""),
        "og_site_name": payload.get("meta", {}).get("og:site_name", ""),
        "meta_title": payload.get("meta", {}).get("title", ""),
    }

    restaurant_name = choose_name(schema, payload.get("title", ""), meta)
    address = guess_address(text, schema)

    website_links = []
    maps_links = []
    social_links = []
    menu_links = []
    ordering_links = []

    for link in links:
        href = link.get("href", "")
        link_text = clean_space(link.get("text", ""))
        combined = f"{href} {link_text}".lower()
        categorized = False

        if any(domain in combined for domain in MAPS_DOMAINS):
            maps_links.append(href)
            categorized = True
        if any(word in combined for word in MENU_WORDS):
            menu_links.append(href)
            categorized = True
        if any(word in combined for word in ORDERING_WORDS):
            ordering_links.append(href)
            categorized = True
        if any(domain in combined for domain in SOCIAL_DOMAINS):
            social_links.append(href)
            categorized = True
        if not categorized and href.startswith(("http://", "https://")):
            website_links.append(href)

    current_url = payload.get("url", "")
    score, reason = score_lead(
        text=text,
        phones=phones,
        address=address,
        menu_links=unique(menu_links),
        ordering_links=unique(ordering_links),
        current_url=current_url,
    )

    return {
        "captured_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "restaurant_name": restaurant_name,
        "phone_numbers": ", ".join(phones),
        "emails": ", ".join(emails),
        "address_guess": address,
        "source_url": current_url,
        "source_title": clean_space(payload.get("title", "")),
        "website_links": "\n".join(unique(website_links[:8])),
        "google_maps_links": "\n".join(unique(maps_links[:5])),
        "social_links": "\n".join(unique(social_links[:8])),
        "menu_links": "\n".join(unique(menu_links[:8])),
        "ordering_links": "\n".join(unique(ordering_links[:8])),
        "has_menu_signal": "yes" if menu_links else "no",
        "has_ordering_signal": "yes" if ordering_links else "no",
        "qrave_fit_score": score,
        "qrave_fit_reason": reason,
        "notes": "",
        "page_text_preview": clean_space(text)[:1200],
    }


def ensure_workbook(path):
    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
        existing_headers = [cell.value for cell in sheet[1]]
        if existing_headers != HEADERS:
            raise RuntimeError(
                f"{path} already exists but its headers do not match this crawler."
            )
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Qrave Leads"
    sheet.append(HEADERS)

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font

    sheet.freeze_panes = "A2"
    return workbook, sheet


def existing_keys(sheet):
    headers = [cell.value for cell in sheet[1]]
    source_url_col = headers.index("source_url") + 1
    phone_col = headers.index("phone_numbers") + 1
    keys = set()
    for row in range(2, sheet.max_row + 1):
        source_url = sheet.cell(row=row, column=source_url_col).value or ""
        phones = sheet.cell(row=row, column=phone_col).value or ""
        keys.add((source_url.strip().lower(), phones.strip()))
    return keys


def append_row(path, data, allow_duplicate=False):
    workbook, sheet = ensure_workbook(path)
    key = (data["source_url"].strip().lower(), data["phone_numbers"].strip())
    if not allow_duplicate and key in existing_keys(sheet):
        return False

    sheet.append([data.get(header, "") for header in HEADERS])

    widths = {
        "A": 20,
        "B": 28,
        "C": 22,
        "D": 26,
        "E": 42,
        "F": 48,
        "G": 36,
        "H": 38,
        "I": 38,
        "J": 38,
        "K": 38,
        "L": 38,
        "M": 16,
        "N": 18,
        "O": 16,
        "P": 44,
        "Q": 28,
        "R": 80,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    workbook.save(path)
    return True


def get_active_page(browser):
    pages = []
    for context in browser.contexts:
        pages.extend(context.pages)
    pages = [page for page in pages if not page.is_closed()]
    if not pages:
        raise RuntimeError("No browser page is open.")
    return pages[-1]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Manual browser crawler for QRave restaurant leads."
    )
    parser.add_argument(
        "--cdp-url",
        default="http://localhost:9222",
        help="Chrome remote debugging URL. Default: http://localhost:9222",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Excel output file. Default: {DEFAULT_OUTPUT}",
    )
    parser.add_argument(
        "--allow-duplicates",
        action="store_true",
        help="Append even when the same URL and phone number were already captured.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    output_path = Path(args.output).resolve()

    with sync_playwright() as playwright:
        try:
            browser = playwright.chromium.connect_over_cdp(args.cdp_url)
        except Error as exc:
            raise SystemExit(
                "Could not connect to Chrome. Start Chrome with remote debugging first.\n"
                "PowerShell command:\n"
                '& "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" '
                '--remote-debugging-port=9222 --user-data-dir="C:\\tmp\\qrave-browser"\n\n'
                f"Original error: {exc}"
            )

        print("Manual QRave lead crawler is running.")
        print("Browse in Chrome, then come back here.")
        print("Press ENTER or type c to capture the current page.")
        print("Type q to quit.")
        print(f"Excel file: {output_path}")

        while True:
            command = input("\nCommand [ENTER/c/q]: ").strip().lower()
            if command == "q":
                break
            if command not in {"", "c"}:
                print("Unknown command. Use ENTER/c/q.")
                continue

            page = get_active_page(browser)
            try:
                data = extract_from_page(page)
                note = input("Optional note for this lead: ").strip()
                data["notes"] = note
                added = append_row(
                    output_path,
                    data,
                    allow_duplicate=args.allow_duplicates,
                )
            except Exception as exc:
                print(f"Capture failed: {exc}")
                continue

            if added:
                print(f"Saved: {data['restaurant_name'] or data['source_title']}")
                print(f"Score: {data['qrave_fit_score']} | {data['qrave_fit_reason']}")
            else:
                print("Skipped duplicate. Use --allow-duplicates if you want every capture.")

        browser.close()
        print(f"Done. Leads saved in: {output_path}")


if __name__ == "__main__":
    main()
